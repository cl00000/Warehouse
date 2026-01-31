# function_create_auxiliary_table.py
"""
创建副表功能模块
"""
import os
import subprocess
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def create_auxiliary_table():
    """
    在桌面创建副表.xlsx文件并自动打开

    :return: (success: bool, message: str)
    """
    try:
        # 获取桌面路径
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        file_path = os.path.join(desktop_path, "副表.xlsx")

        # 检查文件是否已存在
        if os.path.exists(file_path):
            return False, "副表已存在，请先删除或重命名现有文件"

        # 创建工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = "副表数据"

        # 设置表头
        headers = ["排除", "刷单"]
        ws.append(headers)

        # 设置表头样式
        header_fill = PatternFill(start_color="9999FF", end_color="9999FF", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

            # 设置列宽
            ws.column_dimensions[chr(64 + col)].width = 15

        # 保存文件
        wb.save(file_path)

        # 自动打开Excel文件
        if os.name == 'nt':  # Windows系统
            os.startfile(file_path)
        elif os.name == 'posix':  # Mac或Linux系统
            subprocess.run(['open', file_path] if sys.platform == 'darwin' else ['xdg-open', file_path])

        return True, ""

    except Exception as e:
        return False, f"创建副表失败: {str(e)}"