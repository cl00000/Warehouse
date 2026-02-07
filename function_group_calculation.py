import pandas as pd
from collections import defaultdict
import os
from datetime import datetime
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from function_group_sub_table import SubTableHandler


class ExcelContrastProcessor:
    def __init__(self):
        self.desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        self.code_dict = {}
        self.priority_dict = {}
        self.messages = []
        self.today_str = datetime.now().strftime("%y%m%d")
        self.has_unmatched_codes = False
        self.non_today_waves = set()
        self.brush_style = "洗脸巾/其它包数"
        self.COLOR_INFO = "green"
        self.COLOR_WARN = "purple"
        self.COLOR_ERROR = "red"
        self.unmatched_waves = defaultdict(list)
        self.partial_unmatched_waves = defaultdict(list)
        self.channel_colors = {}
        self.order_type_colors = {}

        # 新增：订单类型映射表
        self.channel_type_map = []
        # 新增：用于跟踪未匹配的映射关系
        self.unmatched_mappings = defaultdict(list)

        # 初始化副表处理器
        self.sub_table_handler = SubTableHandler(self.desktop_path)

    def is_empty(self, value):
        if pd.isna(value):
            return True
        if isinstance(value, str) and value.strip() == "":
            return True
        return False

    def load_data(self):
        file1_path = os.path.join(self.desktop_path, "1.xlsx")
        code_file_path = os.path.join(self.desktop_path, "编码对应关系.xlsx")

        try:
            try:
                # 从"编码对应关系.xlsx"文件的"映射"sheet中读取
                map_df = pd.read_excel(code_file_path, sheet_name="映射")
                self.channel_type_map = []

                # 检查必需的列是否存在
                required_map_columns = ["渠道", "类型", "输出渠道", "输出类型"]
                if all(col in map_df.columns for col in required_map_columns):
                    for _, row in map_df.iterrows():
                        if not (pd.isna(row["渠道"]) or pd.isna(row["类型"]) or
                                pd.isna(row["输出渠道"]) or pd.isna(row["输出类型"])):
                            self.channel_type_map.append({
                                "渠道": str(row["渠道"]).strip(),
                                "类型": str(row["类型"]).strip(),
                                "输出渠道": str(row["输出渠道"]).strip(),
                                "输出类型": str(row["输出类型"]).strip()
                            })
                else:
                    missing_cols = [col for col in required_map_columns if col not in map_df.columns]
                    return None, None, f"映射表中缺少必要的列: {', '.join(missing_cols)}"

                if not self.channel_type_map:
                    return None, None, "映射表中没有有效数据"

            except ValueError as e:
                if "Worksheet named '映射' not found" in str(e):
                    return None, None, f"编码对应关系.xlsx文件中没有找到'映射'工作表"
                else:
                    return None, None, f"读取映射表时出错: {str(e)}"
            except Exception as e:
                return None, None, f"读取映射表时出错: {str(e)}"

            df1 = pd.read_excel(file1_path, sheet_name="Sheet1")
            required_columns = ['打印波次', '店铺', '货品商家编码', '订单类型', '打单员']
            missing_columns = [col for col in required_columns if col not in df1.columns]
            if missing_columns:
                return None, None, f"1.xlsx中缺少必要的列: {', '.join(missing_columns)}"

            key_columns = ['打印波次', '店铺', '货品商家编码', '订单类型', '打单员']
            null_rows = df1[df1[key_columns].isnull().any(axis=1)]
            if not null_rows.empty:
                null_row_numbers = [idx + 2 for idx in null_rows.index]
                return None, None, f"发现空值行，行号: {', '.join(map(str, null_row_numbers))}"

            if '打单员' in df1.columns:
                df1 = df1[df1['打单员'].astype(str).str.contains('打单')]

            current_mmdd = datetime.now().strftime("%m%d")
            for idx, row in enumerate(df1.itertuples(index=False)):
                wave_value = getattr(row, '打印波次')
                if not self.is_empty(wave_value):
                    wave_str = str(wave_value).strip()
                    if wave_str.startswith("PB" + datetime.now().strftime("%y")) and len(wave_str) >= 10:
                        wave_mmdd = wave_str[4:8]
                        if wave_mmdd != current_mmdd:
                            self.non_today_waves.add(wave_str)

            if self.non_today_waves:
                self.messages.append({
                    "text": f"发现非当天波次，已忽略: {', '.join(sorted(self.non_today_waves))}",
                    "color": self.COLOR_INFO
                })

            df_code = pd.read_excel(code_file_path, sheet_name="编码")
            df_code["优先级"] = df_code["优先级"].ffill()

            brush_rows = df_code[df_code["货品商家编码"].astype(str).str.strip() == "刷单"]
            if not brush_rows.empty:
                self.brush_style = brush_rows.iloc[0]["名称"]

            try:
                color_df = pd.read_excel(code_file_path, sheet_name="颜色")

                # 渠道颜色
                if "输出渠道" in color_df.columns and "颜色" in color_df.columns:
                    for _, row in color_df.iterrows():
                        if not pd.isna(row["输出渠道"]) and not pd.isna(row["颜色"]):
                            channel = str(row["输出渠道"]).strip()
                            color = str(row["颜色"]).strip().upper()
                            self.channel_colors[channel] = Font(color=color)

                # 订单类型颜色（允许同一个sheet复用）
                if "输出类型" in color_df.columns and "颜色.1" in color_df.columns:
                    for _, row in color_df.iterrows():
                        if not pd.isna(row["输出类型"]) and not pd.isna(row["颜色.1"]):
                            order_type = str(row["输出类型"]).strip()
                            color = str(row["颜色.1"]).strip().upper()
                            self.order_type_colors[order_type] = Font(color=color)

            except ValueError as e:
                if "Worksheet named '颜色' not found" in str(e):
                    return None, None, "编码对应关系.xlsx 中未找到 sheet「颜色」"
                else:
                    return None, None, f"读取颜色配置时出错: {str(e)}"
            except Exception as e:
                return None, None, f"读取颜色配置时出错: {str(e)}"

            # 加载副表数据 - 使用新的处理器（副表保持单独文件）
            sub_table_error = self.sub_table_handler.load_sub_table()
            if sub_table_error:
                return None, None, sub_table_error

            # 添加副表消息
            self.messages.extend(self.sub_table_handler.sub_table_messages)

            return df1, df_code, None

        except Exception as e:
            return None, None, f"读取数据时出错: {str(e)}"

    def build_code_mappings(self, df_code):
        self.style_sort = {}
        for _, row in df_code.iterrows():
            code = str(row["货品商家编码"]).strip()
            name = str(row["名称"]).strip()
            pri = int(row["优先级"]) if not pd.isna(row["优先级"]) else 999
            sort_val = int(row["排序"]) if "排序" in df_code.columns and not pd.isna(row.get("排序")) else 999
            self.code_dict[code] = name
            self.priority_dict[code] = pri
            if name not in self.style_sort or sort_val < self.style_sort[name]:
                self.style_sort[name] = sort_val

    def get_mapped_channel_and_type(self, shop, order_type_tag, wave_str, row_idx):
        """根据店铺和订单类型获取映射后的渠道和类型"""
        shop = str(shop).strip() if not pd.isna(shop) else ""
        order_type_tag = str(order_type_tag).strip() if not pd.isna(order_type_tag) else ""

        # 从店铺名中提取渠道部分（第一个/号前的内容）
        if '/' in shop:
            channel_part = shop.split('/')[0] + "/"
        else:
            channel_part = shop + "/"  # 如果没有斜杠，直接添加斜杠

        # 在映射表中查找匹配项
        for mapping in self.channel_type_map:
            if mapping["渠道"] == channel_part and mapping["类型"] == order_type_tag:
                return mapping["输出渠道"], mapping["输出类型"]

        # 如果没有找到匹配项，记录到未匹配集合中
        mapping_key = f"{channel_part}|{order_type_tag}"
        self.unmatched_mappings[mapping_key].append({
            "wave": wave_str,
            "row_idx": row_idx + 2,  # Excel行号（从1开始，加上标题行）
            "shop": shop,
            "order_type": order_type_tag
        })

        # 返回None表示没有匹配项
        return None, None

    def parse_codes_and_quantities(self, product_code):
        # 解析“货品商家编码”字段，支持复合写法，如 "A*2;B*3;C"
        codes = []
        quantities = {}
        parts = [p.strip() for p in str(product_code).split(';') if p.strip()]

        for part in parts:
            if '*' in part:
                code_part, _, qty_part = part.partition('*')
                try:
                    qty = int(qty_part.strip())
                    code = code_part.strip()
                    codes.append(code)
                    quantities[code] = qty
                except ValueError:
                    codes.append(part)
                    quantities[part] = 1
            else:
                codes.append(part)
                quantities[part] = 1

        return codes, quantities

    class UniformSelector:
        # 用于在同一波次 / 同一组合下，实现「公平、均匀」的款式选择策略，避免总是选同一个编码。
        def __init__(self):
            self.group_state = defaultdict(lambda: {
                'index': 0,
                'sorted_codes': None,
                'counters': defaultdict(int),
                '_index': 0,
                'max_qty_codes': None
            })

        def select(self, candidates, quantities, priority_dict, group_key):
            if not candidates:
                return None

            state = self.group_state[group_key]
            priorities = {priority_dict[code] for code in candidates}
            same_priority = len(priorities) == 1
            qty_values = {quantities[code] for code in candidates}
            same_quantity = len(qty_values) == 1

            if same_priority and same_quantity:
                if state['sorted_codes'] is None:
                    state['sorted_codes'] = sorted(candidates)
                selected_code = state['sorted_codes'][state['index'] % len(state['sorted_codes'])]
                state['index'] += 1
                state['counters'][selected_code] += 1
                return selected_code

            elif same_priority:
                if state['max_qty_codes'] is None:
                    max_qty = max(quantities[code] for code in candidates)
                    state['max_qty_codes'] = [code for code in candidates if quantities[code] == max_qty]

                if len(state['max_qty_codes']) > 1:
                    selected_code = state['max_qty_codes'][state['_index'] % len(state['max_qty_codes'])]
                    state['_index'] += 1
                else:
                    selected_code = state['max_qty_codes'][0]
                state['counters'][selected_code] += 1
                return selected_code

            else:
                min_priority = min(priority_dict[code] for code in candidates)
                priority_candidates = [code for code in candidates if priority_dict[code] == min_priority]
                return self.select(priority_candidates, quantities, priority_dict, group_key)

    def process_data(self, df1):
        result_data = []
        selector = self.UniformSelector()
        excluded_row_count = 0
        sub_handler = self.sub_table_handler
        get_mapping = self.get_mapped_channel_and_type
        wave_stats = {
            'brush_waves': defaultdict(list)
        }
        # 未匹配映射的计数器
        unmapped_row_count = 0

        max_sort = max(self.style_sort.values()) if self.style_sort else 999

        # 使用 defaultdict 来收集数据
        group_data = defaultdict(lambda: defaultdict(int))

        for idx, row in df1.iterrows():
            wave_value = row['打印波次']
            wave_str = str(wave_value).strip() if not self.is_empty(wave_value) else None
            shop = str(row.get("店铺", "")).strip()
            order_type_tag = str(row.get("订单类型", "")).strip()
            product_code = row["货品商家编码"]

            if wave_str and wave_str in self.non_today_waves:
                continue

            # 使用映射表获取输出渠道和输出类型
            output_channel, output_type = get_mapping(
                shop, order_type_tag, wave_str, idx
            )

            # 如果映射返回None，跳过该行
            if output_channel is None or output_type is None:
                unmapped_row_count += 1
                continue

            # 使用副表处理器检查排除波次
            if wave_str and sub_handler.is_excluded_wave(wave_str):
                excluded_row_count += 1
                continue

            # 使用副表处理器检查刷单波次
            if wave_str and self.sub_table_handler.is_brush_wave(wave_str):
                wave_stats['brush_waves'][wave_str].append(idx)
                style = self.brush_style
                key = (output_channel, output_type, style)
                group_data[key]['单量'] += 1
                group_data[key]['实际数量'] += 1
            else:
                code_list, quantities = self.parse_codes_and_quantities(product_code)
                if not code_list:
                    self.unmatched_waves[wave_str].append((idx + 2, []))  # 空编码视为未匹配
                    continue

                unmatched_codes = [code for code in code_list if code not in self.code_dict]
                if unmatched_codes:
                    self.has_unmatched_codes = True
                    if len(unmatched_codes) == len(code_list):
                        self.unmatched_waves[wave_str].append((idx + 2, unmatched_codes))
                        continue
                    else:
                        self.partial_unmatched_waves[wave_str].append((idx + 2, unmatched_codes))

                valid_codes = [code for code in code_list if code in self.code_dict]
                if not valid_codes:
                    continue

                # 为每个有效编码累加实际数量
                for code in valid_codes:
                    style = self.code_dict[code]
                    key = (output_channel, output_type, style)
                    group_data[key]['实际数量'] += quantities[code]

                # 选择主编码并累加单量
                min_pri = min(self.priority_dict[c] for c in valid_codes)
                candidates = [c for c in valid_codes if self.priority_dict[c] == min_pri]
                group_key = tuple(sorted(candidates))
                main_code = selector.select(candidates, quantities, self.priority_dict, group_key)
                if main_code:
                    main_style = self.code_dict[main_code]
                    key = (output_channel, output_type, main_style)
                    group_data[key]['单量'] += 1

        # 检查未匹配的映射关系
        if self.unmatched_mappings:
            warning_details = []

            # 按渠道分组显示未匹配的行
            for mapping_key, rows in self.unmatched_mappings.items():
                channel, order_type = mapping_key.split("|")
                row_numbers = sorted([r["row_idx"] for r in rows])

                # 只显示前10行，避免消息过长
                if len(row_numbers) > 10:
                    row_display = f"{', '.join(map(str, row_numbers[:10]))}, ...等{len(row_numbers)}行"
                else:
                    row_display = ', '.join(map(str, row_numbers))

                warning_details.append(f"渠道: {channel}, 类型: {order_type}, 行号: {row_display}")

            # 用竖线分隔所有条目
            warning_msg = f"警告：发现 {unmapped_row_count} 行数据没有匹配的订单类型组合： | " + ' | '.join(
                warning_details)

            self.messages.append({
                "text": warning_msg.strip(),
                "color": self.COLOR_ERROR
            })

        if self.unmatched_waves:
            # 汇总所有未匹配的行和编码
            all_unmatched = []
            for items in self.unmatched_waves.values():
                all_unmatched.extend(items)

            all_unmatched_sorted = sorted(all_unmatched, key=lambda x: x[0])
            unmatched_count = len(all_unmatched_sorted)

            details = []
            for row, codes in all_unmatched_sorted[:20]:
                codes_str = ', '.join(codes) if codes else '无编码'
                details.append(f"行{row}: {codes_str}")

            if unmatched_count > 20:
                details.append(f"等共 {unmatched_count} 行")

            # 使用竖线分隔
            details_str = ' | '.join(details)

            self.messages.append({
                "text": f"发现 {unmatched_count} 行数据未匹配到任何编码，详情：{details_str}",
                "color": self.COLOR_ERROR
            })

        if self.partial_unmatched_waves:
            # 汇总所有部分未匹配的行和编码
            all_partial = []
            for items in self.partial_unmatched_waves.values():
                all_partial.extend(items)

            all_partial_sorted = sorted(all_partial, key=lambda x: x[0])
            partial_count = len(all_partial_sorted)

            details = []
            for row, codes in all_partial_sorted[:20]:
                codes_str = ', '.join(codes)
                details.append(f"行{row}: {codes_str}")

            if partial_count > 20:
                details.append(f"等共 {partial_count} 行")

            # 使用竖线分隔
            details_str = ' | '.join(details)

            self.messages.append({
                "text": f"发现 {partial_count} 行数据部分编码未匹配，详情：{details_str}",
                "color": self.COLOR_ERROR
            })

        if wave_stats['brush_waves']:
            brush_waves = sorted(wave_stats['brush_waves'].keys())
            brush_count = sum(len(rows) for rows in wave_stats['brush_waves'].values())
            self.messages.append({
                "text": f"已处理 {brush_count} 条刷单波次:{'，'.join(brush_waves)}",
                "color": self.COLOR_INFO
            })

        # 获取副表处理器中的已处理排除波次
        if self.sub_table_handler.processed_excluded_waves:
            excluded_waves_sorted = sorted(self.sub_table_handler.processed_excluded_waves)
            wave_details = [f"{wave}" for wave in excluded_waves_sorted]
            self.messages.append({
                "text": f"已排除 {excluded_row_count} 条数据，来自波次: {', '.join(wave_details)}",
                "color": self.COLOR_INFO
            })

        # 获取副表警告信息并添加到主消息中
        sub_table_warnings = self.sub_table_handler.get_warnings()
        self.messages.extend(sub_table_warnings)

        # 创建优先级映射（按字母顺序排序）
        channel_priority = {"自营": 0, "分销": 1, "代发": 2}
        # 订单类型优先级
        order_type_priority = {"新订单": 0, "补发单": 1, "批采单": 2}

        # 收集只有单量 > 0 的组
        group_list = []
        for (channel, otype, style), data in group_data.items():
            if data['单量'] > 0:
                group_list.append({
                    "订单渠道": channel,
                    "订单类型": otype,
                    "款式": style,
                    "单量": data['单量'],
                    "实际数量": data['实际数量'],
                    "订单渠道优先级": channel_priority.get(channel, 999),
                    "订单类型优先级": order_type_priority.get(otype, 999),
                    "款式排序": self.style_sort.get(style, max_sort + 1)
                })

        if group_list:
            group_counts = pd.DataFrame(group_list)
            # 按照订单渠道优先级、订单类型优先级和款式排序值升序排列
            group_counts = group_counts.sort_values(["订单渠道优先级", "订单类型优先级", "款式排序"])
            # 删除临时列
            group_counts = group_counts.drop(columns=["订单渠道优先级", "订单类型优先级", "款式排序"])

            for _, row in group_counts.iterrows():
                result_data.append([
                    row["订单渠道"],
                    row["订单类型"],
                    row["款式"],
                    row["单量"],
                    row["实际数量"],
                    row["款式"] == "（未匹配到编码）",
                    row["款式"] == "空值"
                ])

        return {
            "data": result_data,
            "messages": self.messages
        }

    def create_output_excel(self, result_data):
        # 根据处理结果生成最终Excel文件：
        wb = Workbook()
        ws = wb.active
        ws.title = "汇总结果"
        header_font = Font(color="FFFFFF", bold=True)
        header_fill = PatternFill(start_color="9999FF", end_color="9999FF", fill_type="solid")
        headers = ["订单渠道", "订单类型", "款式", "单量", "实际数量"]
        ws.append(headers)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill

        result_df = pd.DataFrame(result_data, columns=[
            "订单渠道", "订单类型", "款式", "单量", "实际数量",
            "未匹配标记", "空值行标记"
        ])

        color_map = {
            "green": Font(color="00AA00"),
            "purple": Font(color="800080"),
            "red": Font(color="FF0000")
        }

        ws.delete_rows(2, ws.max_row)

        max_main_rows = len(result_df) if result_df is not None else 0

        for i in range(max_main_rows):
            actual_row = i + 2
            main_idx = i
            main_row = [
                result_df.iloc[main_idx]["订单渠道"],
                result_df.iloc[main_idx]["订单类型"],
                result_df.iloc[main_idx]["款式"],
                result_df.iloc[main_idx]["单量"],
                result_df.iloc[main_idx]["实际数量"]
            ]

            for col, val in enumerate(main_row, start=1):
                ws.cell(row=actual_row, column=col, value=val)

            if result_df.iloc[main_idx]["空值行标记"]:
                for col in range(3, 6):
                    ws.cell(row=actual_row, column=col).font = color_map["red"]
            elif result_df.iloc[main_idx]["未匹配标记"] or result_df.iloc[main_idx]["款式"] == "（未匹配到部分编码）":
                for col in range(3, 6):
                    ws.cell(row=actual_row, column=col).font = color_map["red"]

            cell_a = ws.cell(row=actual_row, column=1)
            val_a = cell_a.value
            if val_a in self.channel_colors:
                cell_a.font = self.channel_colors[val_a]

            cell_b = ws.cell(row=actual_row, column=2)
            val_b = cell_b.value
            if val_b in self.order_type_colors:
                cell_b.font = self.order_type_colors[val_b]

        ws.append([])

        if self.messages:
            for msg in self.messages:
                if isinstance(msg, dict):
                    ws.append([msg["text"]])
                    last_row = ws.max_row
                    cell = ws.cell(row=last_row, column=1)
                    cell.font = color_map.get(msg["color"], Font())
                else:
                    ws.append([msg])

        # 设置列宽
        column_widths = {'A': 15, 'B': 15, 'C': 30, 'D': 10, 'E': 15}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        output_path = os.path.join(self.desktop_path, "输出结果.xlsx")
        wb.save(output_path)
        return output_path

    def open_file_windows(self, file_path):
        try:
            os.startfile(file_path)
            return True
        except Exception as e:
            print(f"自动打开文件失败: {str(e)}")
            return False

    def process(self):
        start_time = time.time()

        try:
            df1, df_code, error = self.load_data()
            if error:
                return False, error

            self.build_code_mappings(df_code)

            result = self.process_data(df1)
            result_data = result["data"]

            elapsed_time = time.time() - start_time
            self.messages.insert(0, {
                "text": f"计算用时：{elapsed_time:.2f} 秒",
                "color": self.COLOR_INFO
            })

            output_path = self.create_output_excel(result_data)
            self.open_file_windows(output_path)

            return True, "处理完成，已自动打开输出文件"


        except Exception as e:
            return False, f"处理过程中出错: {str(e)}"


def group_calculation():
    processor = ExcelContrastProcessor()
    success, message = processor.process()

    return success, message


